#!/usr/bin/env python3
"""
Script para importar datos de jornales desde archivos Excel
(ADELANTO + HS. EXTRAS 2025.xlsx y 2026.xlsx)

Uso:
    python scripts/importar_jornales_excel.py --analizar  # Solo muestra estructura
    python scripts/importar_jornales_excel.py --importar  # Importa los datos
    python scripts/importar_jornales_excel.py --importar --dry-run  # Simula sin guardar

Estructura esperada del Excel:
- Hojas mensuales: ENERO, FEBRERO, ..., DICIEMBRE
- Fila 1: Encabezados con días del mes
- Columna A: Nombre del empleado
- Columna B: VALOR HORA EXTRA
- Columnas C+: Días del mes con subcolumnas ADELANTO, HS, EXTRAS
"""

import os
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
import argparse
from typing import Optional

# Agregar el directorio padre al path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import select

# Importar modelos y configuración
from app.db.base import SessionLocal, engine
from app.models.empleado import Empleado, MovimientoNomina
from app.models.usuario import Usuario

# UUID del admin para importación
ADMIN_USER_ID = '12b80de2-f104-4da3-872f-145bbcc49a3b'

# Mapeo de nombres de hojas a número de mes
MESES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
}

# Nombres de empleados en Excel -> nombres en BD (mapeo manual si difieren)
MAPEO_EMPLEADOS = {
    # "NOMBRE EN EXCEL": "nombre en BD",
    # Agregar aquí si hay diferencias
}


def get_semana_del_mes(fecha: date) -> int:
    """Calcula la semana del mes (1-5)"""
    dia = fecha.day
    if dia <= 7:
        return 1
    elif dia <= 14:
        return 2
    elif dia <= 21:
        return 3
    elif dia <= 28:
        return 4
    else:
        return 5


def normalizar_nombre(nombre: str) -> str:
    """Normaliza el nombre para comparación"""
    if not nombre:
        return ""
    return nombre.strip().upper()


def buscar_empleado(db: Session, nombre_excel: str) -> Optional[Empleado]:
    """Busca un empleado por nombre (flexible)"""
    nombre_norm = normalizar_nombre(nombre_excel)

    # Verificar si hay mapeo manual
    if nombre_norm in MAPEO_EMPLEADOS:
        nombre_buscar = MAPEO_EMPLEADOS[nombre_norm]
    else:
        nombre_buscar = nombre_norm

    # Buscar por nombre completo
    empleados = db.execute(
        select(Empleado).where(Empleado.activo == True)
    ).scalars().all()

    for emp in empleados:
        nombre_completo = normalizar_nombre(emp.nombre_completo)
        nombre_apellido = normalizar_nombre(f"{emp.nombre} {emp.apellido}")
        apellido_nombre = normalizar_nombre(f"{emp.apellido} {emp.nombre}")

        if nombre_buscar in [nombre_completo, nombre_apellido, apellido_nombre]:
            return emp

        # Búsqueda parcial
        if nombre_buscar in nombre_completo or nombre_completo in nombre_buscar:
            return emp

    return None


def analizar_excel(filepath: str) -> dict:
    """Analiza la estructura del archivo Excel"""
    print(f"\n{'='*60}")
    print(f"Analizando: {filepath}")
    print('='*60)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    resultado = {
        'archivo': filepath,
        'hojas': [],
        'empleados_encontrados': set(),
        'estructura': {}
    }

    for sheet_name in wb.sheetnames:
        mes_num = MESES.get(sheet_name.upper())
        if not mes_num:
            print(f"  ⚠️  Hoja ignorada: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"\n📅 Hoja: {sheet_name} (Mes {mes_num})")
        resultado['hojas'].append(sheet_name)

        # Analizar estructura de columnas (fila 1)
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append((col, str(val)))

        print(f"   Columnas detectadas: {len(headers)}")

        # Analizar empleados (columna A, desde fila 2)
        empleados_hoja = []
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row=row, column=1).value
            if nombre and isinstance(nombre, str) and nombre.strip():
                nombre_clean = nombre.strip()
                # Ignorar filas de totales
                if 'TOTAL' not in nombre_clean.upper() and 'SEMANA' not in nombre_clean.upper():
                    empleados_hoja.append(nombre_clean)
                    resultado['empleados_encontrados'].add(nombre_clean)

        print(f"   Empleados encontrados: {len(empleados_hoja)}")
        for emp in empleados_hoja:
            print(f"      - {emp}")

        # Analizar estructura de días
        # Buscar columnas que contengan números de días
        dias_cols = {}
        for col in range(1, min(ws.max_column + 1, 100)):  # Limitar búsqueda
            header = ws.cell(row=1, column=col).value
            if header and isinstance(header, (int, float)):
                dia = int(header)
                if 1 <= dia <= 31:
                    # Verificar subcolumnas (ADELANTO, HS, EXTRAS)
                    dias_cols[dia] = col

        print(f"   Días con datos: {sorted(dias_cols.keys())}")
        resultado['estructura'][sheet_name] = {
            'empleados': empleados_hoja,
            'dias': dias_cols
        }

    wb.close()
    return resultado


def importar_excel(filepath: str, anio: int, db: Session, dry_run: bool = False) -> dict:
    """Importa datos de un archivo Excel"""
    print(f"\n{'='*60}")
    print(f"Importando: {filepath} (Año {anio})")
    print(f"Modo: {'SIMULACIÓN' if dry_run else 'REAL'}")
    print('='*60)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    stats = {
        'adelantos': 0,
        'horas_extras': 0,
        'errores': [],
        'empleados_no_encontrados': set(),
        'valores_hora_actualizados': 0
    }

    for sheet_name in wb.sheetnames:
        mes_num = MESES.get(sheet_name.upper())
        if not mes_num:
            continue

        ws = wb[sheet_name]
        print(f"\n📅 Procesando: {sheet_name} ({mes_num}/{anio})")

        # Detectar estructura de columnas
        # La estructura típica es:
        # A: Nombre | B: VALOR HORA | C-onwards: días con ADELANTO, HS, EXTRAS

        # Buscar la fila de encabezado con los días
        header_row = 1
        dia_columnas = {}  # {dia: {'adelanto': col, 'hs': col, 'extras': col}}

        # Buscar columnas de días (números del 1 al 31)
        current_dia = None
        for col in range(3, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value

            if cell_val is not None:
                if isinstance(cell_val, (int, float)) and 1 <= int(cell_val) <= 31:
                    current_dia = int(cell_val)
                    dia_columnas[current_dia] = {'col_inicio': col}
                elif isinstance(cell_val, str):
                    cell_upper = cell_val.upper().strip()
                    if current_dia and current_dia in dia_columnas:
                        if 'ADELANTO' in cell_upper:
                            dia_columnas[current_dia]['adelanto'] = col
                        elif cell_upper == 'HS' or 'HORA' in cell_upper:
                            dia_columnas[current_dia]['hs'] = col
                        elif 'EXTRA' in cell_upper:
                            dia_columnas[current_dia]['extras'] = col

        # Si no encontramos estructura clara, intentar estructura alternativa
        # donde cada día tiene 3 columnas consecutivas: ADELANTO, HS, EXTRAS
        if not any('adelanto' in d for d in dia_columnas.values()):
            print("   Intentando estructura alternativa (3 cols por día)...")
            dia_columnas = {}
            col = 3  # Empezar después de nombre y valor hora
            for dia in range(1, 32):
                if col + 2 <= ws.max_column:
                    dia_columnas[dia] = {
                        'adelanto': col,
                        'hs': col + 1,
                        'extras': col + 2
                    }
                    col += 3

        print(f"   Días mapeados: {len(dia_columnas)}")

        # Procesar empleados (desde fila 2)
        for row in range(2, ws.max_row + 1):
            nombre_excel = ws.cell(row=row, column=1).value

            if not nombre_excel or not isinstance(nombre_excel, str):
                continue

            nombre_excel = nombre_excel.strip()

            # Ignorar filas de totales
            if any(x in nombre_excel.upper() for x in ['TOTAL', 'SEMANA', 'MES']):
                continue

            # Buscar empleado en BD
            empleado = buscar_empleado(db, nombre_excel)

            if not empleado:
                stats['empleados_no_encontrados'].add(nombre_excel)
                continue

            print(f"\n   👤 {nombre_excel} -> {empleado.nombre_completo}")

            # Obtener valor hora extra del Excel (columna B)
            valor_hora_excel = ws.cell(row=row, column=2).value
            if valor_hora_excel and isinstance(valor_hora_excel, (int, float)) and valor_hora_excel > 0:
                valor_hora = Decimal(str(valor_hora_excel))
                if not dry_run and empleado.valor_hora_extra != valor_hora:
                    empleado.valor_hora_extra = valor_hora
                    stats['valores_hora_actualizados'] += 1
                    print(f"      💰 Valor hora extra actualizado: ${valor_hora}")

            # Procesar cada día
            for dia, cols in dia_columnas.items():
                try:
                    fecha = date(anio, mes_num, dia)
                except ValueError:
                    continue  # Día inválido para el mes

                semana = get_semana_del_mes(fecha)

                # Procesar ADELANTO
                if 'adelanto' in cols:
                    adelanto_val = ws.cell(row=row, column=cols['adelanto']).value
                    if adelanto_val and isinstance(adelanto_val, (int, float)) and adelanto_val > 0:
                        monto = Decimal(str(adelanto_val))

                        if not dry_run:
                            # Verificar si ya existe
                            existente = db.execute(
                                select(MovimientoNomina).where(
                                    MovimientoNomina.empleado_id == empleado.id,
                                    MovimientoNomina.fecha == fecha,
                                    MovimientoNomina.tipo == 'adelanto'
                                )
                            ).scalar_one_or_none()

                            if not existente:
                                from uuid import UUID
                                mov = MovimientoNomina(
                                    empleado_id=empleado.id,
                                    tipo='adelanto',
                                    concepto='Adelanto de sueldo',
                                    descripcion=f'Importado desde Excel {anio}',
                                    periodo_mes=mes_num,
                                    periodo_anio=anio,
                                    monto=monto,
                                    es_debito=True,
                                    fecha=fecha,
                                    semana=semana,
                                    pagado=True,
                                    fecha_pago=fecha,
                                    registrado_por_id=UUID(ADMIN_USER_ID)
                                )
                                db.add(mov)
                                stats['adelantos'] += 1
                        else:
                            stats['adelantos'] += 1

                        print(f"      📅 {fecha}: Adelanto ${monto}")

                # Procesar HORAS EXTRAS
                if 'hs' in cols:
                    hs_val = ws.cell(row=row, column=cols['hs']).value
                    if hs_val and isinstance(hs_val, (int, float)) and hs_val > 0:
                        cantidad_horas = Decimal(str(hs_val))

                        # Obtener valor hora (del empleado o del Excel)
                        valor_hora = empleado.valor_hora_extra or Decimal('0')
                        if valor_hora_excel and isinstance(valor_hora_excel, (int, float)):
                            valor_hora = Decimal(str(valor_hora_excel))

                        monto = cantidad_horas * valor_hora

                        if not dry_run:
                            existente = db.execute(
                                select(MovimientoNomina).where(
                                    MovimientoNomina.empleado_id == empleado.id,
                                    MovimientoNomina.fecha == fecha,
                                    MovimientoNomina.tipo == 'hora_extra'
                                )
                            ).scalar_one_or_none()

                            if not existente:
                                from uuid import UUID
                                mov = MovimientoNomina(
                                    empleado_id=empleado.id,
                                    tipo='hora_extra',
                                    concepto='Horas extras',
                                    descripcion=f'Importado desde Excel {anio}',
                                    periodo_mes=mes_num,
                                    periodo_anio=anio,
                                    monto=monto,
                                    es_debito=False,
                                    fecha=fecha,
                                    semana=semana,
                                    cantidad_horas=cantidad_horas,
                                    valor_hora=valor_hora,
                                    pagado=True,
                                    fecha_pago=fecha,
                                    registrado_por_id=UUID(ADMIN_USER_ID)
                                )
                                db.add(mov)
                                stats['horas_extras'] += 1
                        else:
                            stats['horas_extras'] += 1

                        print(f"      📅 {fecha}: {cantidad_horas} HS x ${valor_hora} = ${monto}")

        if not dry_run:
            db.commit()

    wb.close()
    return stats


def main():
    parser = argparse.ArgumentParser(description='Importar jornales desde Excel')
    parser.add_argument('--analizar', action='store_true', help='Solo analizar estructura')
    parser.add_argument('--importar', action='store_true', help='Importar datos')
    parser.add_argument('--dry-run', action='store_true', help='Simular sin guardar')
    parser.add_argument('--archivo', type=str, help='Archivo específico a procesar')
    args = parser.parse_args()

    # Rutas de archivos
    base_dir = Path(__file__).resolve().parent.parent.parent
    archivos = [
        (base_dir / 'ADELANTO + HS. EXTRAS 2025.xlsx', 2025),
        (base_dir / 'ADELANTO + HS. EXTRAS 2026.xlsx', 2026),
    ]

    if args.archivo:
        archivos = [(Path(args.archivo), int(input("Año del archivo: ")))]

    # Filtrar archivos que existen
    archivos = [(f, a) for f, a in archivos if f.exists()]

    if not archivos:
        print("❌ No se encontraron archivos Excel para procesar")
        print(f"   Buscando en: {base_dir}")
        return

    print(f"📁 Archivos encontrados: {len(archivos)}")
    for f, a in archivos:
        print(f"   - {f.name} (Año {a})")

    if args.analizar:
        for filepath, anio in archivos:
            analizar_excel(str(filepath))
        return

    if args.importar:
        db = SessionLocal()
        try:
            # Mostrar empleados existentes
            empleados = db.execute(
                select(Empleado).where(Empleado.activo == True)
            ).scalars().all()

            print(f"\n👥 Empleados en BD ({len(empleados)}):")
            for emp in empleados:
                print(f"   - {emp.nombre_completo} (Valor HS Extra: ${emp.valor_hora_extra or 0})")

            total_stats = {
                'adelantos': 0,
                'horas_extras': 0,
                'valores_hora_actualizados': 0,
                'empleados_no_encontrados': set()
            }

            for filepath, anio in archivos:
                stats = importar_excel(str(filepath), anio, db, dry_run=args.dry_run)
                total_stats['adelantos'] += stats['adelantos']
                total_stats['horas_extras'] += stats['horas_extras']
                total_stats['valores_hora_actualizados'] += stats['valores_hora_actualizados']
                total_stats['empleados_no_encontrados'].update(stats['empleados_no_encontrados'])

            print(f"\n{'='*60}")
            print("📊 RESUMEN FINAL")
            print('='*60)
            print(f"   ✅ Adelantos importados: {total_stats['adelantos']}")
            print(f"   ✅ Horas extras importadas: {total_stats['horas_extras']}")
            print(f"   ✅ Valores hora actualizados: {total_stats['valores_hora_actualizados']}")

            if total_stats['empleados_no_encontrados']:
                print(f"\n   ⚠️  Empleados NO encontrados en BD:")
                for emp in sorted(total_stats['empleados_no_encontrados']):
                    print(f"      - {emp}")
                print("\n   💡 Agrega estos empleados a la BD o actualiza MAPEO_EMPLEADOS")

            if args.dry_run:
                print(f"\n   ℹ️  Modo DRY-RUN: No se guardaron cambios")
            else:
                print(f"\n   ✅ Datos guardados correctamente")

        except Exception as e:
            print(f"\n❌ Error: {e}")
            import traceback
            traceback.print_exc()
            db.rollback()
        finally:
            db.close()
        return

    # Si no se especifica acción, mostrar ayuda
    parser.print_help()


if __name__ == '__main__':
    main()
